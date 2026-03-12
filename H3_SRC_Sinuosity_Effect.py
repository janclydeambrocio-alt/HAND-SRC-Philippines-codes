"""
──────────────────────
This script performs SRC Generation with varying Manning's n roughness coefficients
for a dual-zone (channel + floodplain) hydraulic model using QGIS and HAND-based
inundation mapping.

It loops through a range of water level multipliers (stage factors), computes
cross-sectional hydraulic properties per pixel, then calculates discharge (Q)
for all possible combinations of Manning's n values per land cover class.

Results are exported to individual Excel files per stage and optionally written
into an existing summary workbook.

Requirements:
    - QGIS Python environment (qgis.core, processing)
    - NumPy, Pandas, OpenPyXL, itertools, math

Author: [Your Name]
Date: [Date]
"""

# ────────────────────────
# IMPORTS
# ────────────────────────
import processing
from qgis.core import (
    QgsProject,
    QgsVectorLayer,
    QgsSpatialIndex
)
import os
import math
import itertools
import numpy as np
import pandas as pd
import openpyxl


# ────────────────────────
# USER CONFIGURATION
# Edit these paths and parameters before running the script.
# ────────────────────────

# Directory where outputs (flood rasters, Excel files) will be saved
base_dir = r"C:\Users\acer\Downloads\aQGIS Content\Official Sample\CDO_NewMethodology\Test Web Length\Output"

# Hydraulic reach length (meters)
reach_length = 3984.6

# DEM pixel resolution (meters)
pixel_res = 5.0

# Valley slope calculated from DEM elevation difference over horizontal distance
valley_S = (42.80544281 - 29.117033) / 2749.334

# Path to the Land Cover shapefile
LC_SHP_PATH = r"C:/Users/acer/Downloads/aQGIS Content/Official Sample/CDO_NewMethodology/Test Web Length/Input 2/LandCover_Vector_3984,6m.shp"

# Path to the channel mask shapefile (used to separate channel vs. floodplain pixels)
CHANNEL_MASK_PATH = r"C:/Users/acer/Downloads/aQGIS Content/Official Sample/CDO_NewMethodology/Test Web Length/FINAL CODE/Smaller Channel/SmallerChannel_Vector.shp"

# Path to the existing Excel workbook where summary results will be written
EXISTING_EXCEL_PATH = r"C:\Users\acer\Downloads\aQGIS Content\Official Sample\CDO_NewMethodology\Test Web Length\QGIS-Calculations-San-Isidro.xlsx"

# Sheet name inside the existing workbook
SUMMARY_SHEET_NAME = "Cabula Curve List"

# Names of the QGIS layers loaded in the project
base_water_layer_name  = "PA_WL_Cabula_10_FINAL"
hand_raster_name       = "HAND_EleveAbvStream_BestSubcatchmentClipped"
slope_raster_name      = "Slope_Degrees_NewDEM"

# Minimum inundation depth (m) for a floodplain pixel to be included in calculations
STABILITY_DEPTH_LIMIT = 0.3

# Critical water depth (m) below which floodplain discharge is assumed zero
H_crit = 1.8

# Sinuosity index — relates valley slope to energy slope
# A higher value means more channel curvature and energy loss
SI = 1.45

# Energy slope, adjusted by sinuosity
Se_adj = valley_S / SI


# ────────────────────────
# MANNING'S n CONFIGURATION PER LAND COVER CLASS
# ────────────────────────
# Each land cover ID maps to a name and a range [Min, Max, Step] for:
#   'chan' = Manning's n for the main channel
#   'fp'   = Manning's n for the floodplain
#
# These ranges are based on Reference Roughness Literature (RRL) values.
# The same range is used for both channel and floodplain in this configuration.

n_configs = {
    1:  {'name': 'Water',      'chan': [0.025, 0.080, 0.005], 'fp': [0.025, 0.080, 0.005]},
    11: {'name': 'Rangelands', 'chan': [0.030, 0.070, 0.001], 'fp': [0.020, 0.040, 0.001]},
    7:  {'name': 'Built Area', 'chan': [0.065, 0.120, 0.050], 'fp': [0.065, 0.120, 0.050]},
    5:  {'name': 'Crops',      'chan': [0.025, 0.055, 0.010], 'fp': [0.025, 0.055, 0.010]},
    2:  {'name': 'Trees',      'chan': [0.120, 0.300, 0.005], 'fp': [0.120, 0.300, 0.005]}
}


# ────────────────────────
# SETUP: CREATE OUTPUT FOLDERS
# ────────────────────────

output_paths = {
    "flood_rasters": os.path.join(base_dir, "Flood Rasters"),
    "excel_files":   os.path.join(base_dir, "Excel Files")
}

for folder in output_paths.values():
    os.makedirs(folder, exist_ok=True)


# ────────────────────────
# HELPER FUNCTION: GENERATE EVENLY SPACED n VALUES
# ────────────────────────

def generate_n_values(start, end, step):
    """
    Returns an array of Manning's n values from 'start' to 'end',
    spaced by 'step'. Both endpoints are included.

    Example:
        generate_n_values(0.025, 0.080, 0.005)
        -> array([0.025, 0.030, 0.035, ..., 0.080])
    """
    count = int(round((end - start) / step)) + 1
    return np.linspace(start, end, count)


# ────────────────────────
# PRE-PROCESSING: BUILD n VALUE COMBINATIONS ACROSS ALL LAND COVER CLASSES
# ────────────────────────

lc_ids = list(n_configs.keys())
paired_n_values = []

for lc_id in lc_ids:
    config = n_configs[lc_id]

    # Generate the list of channel and floodplain n values for this land cover
    channel_n_values    = generate_n_values(*config['chan'])
    floodplain_n_values = generate_n_values(*config['fp'])

    # Pair each channel value with its corresponding floodplain value
    # Note: If channel and floodplain ranges have different lengths,
    #       zip() stops at the shorter list.
    n_pairs = list(zip(
        np.round(channel_n_values, 3),
        np.round(floodplain_n_values, 3)
    ))

    paired_n_values.append(n_pairs)

# Generate all possible combinations across all land cover classes
all_n_combinations = list(itertools.product(*paired_n_values))

print(f"Total n combinations to test: {len(all_n_combinations)}")


# ────────────────────────
# LOAD QGIS LAYERS
# ────────────────────────

def get_layer_by_name(layer_name):
    """Returns a QGIS layer from the current project by its name."""
    matches = QgsProject.instance().mapLayersByName(layer_name)
    if not matches:
        raise ValueError(f"Layer not found in project: '{layer_name}'")
    return matches[0]


base_water_layer = get_layer_by_name(base_water_layer_name)
hand_layer       = get_layer_by_name(hand_raster_name)
slope_layer      = get_layer_by_name(slope_raster_name)

# Load the channel mask and build a spatial index for fast lookup
channel_mask_layer    = QgsVectorLayer(CHANNEL_MASK_PATH, "Channel_Mask", "ogr")
channel_spatial_index = QgsSpatialIndex(channel_mask_layer.getFeatures())


# ────────────────────────
# MAIN PROCESSING LOOP: ITERATE OVER STAGE FACTORS (0.1 to 3.0)
# ────────────────────────

geometry_log = []  # Stores summary info for each stage (area, min Q, etc.)

for stage_index in range(1, 31):

    # Current stage factor (e.g., 0.1, 0.2, ..., 3.0)
    stage_factor = stage_index / 10.0
    stage_label  = f"{stage_factor:.1f}"

    print(f"\nProcessing stage factor: {stage_label}m ...")

    # ------------------------------------------------------------------
    # STEP 1: Generate inundation depth raster using raster calculator
    # Formula: (Water Level * Stage Factor) - HAND > 0 gives flood depth
    # ------------------------------------------------------------------
    inundation_raster_path = os.path.join(
        output_paths["flood_rasters"],
        f"Inun_{stage_label}.tif"
    )

    inundation_expression = (
        f'(("{base_water_layer_name}@1" * {stage_factor}) - "{hand_raster_name}@1" > 0) * '
        f'(("{base_water_layer_name}@1" * {stage_factor}) - "{hand_raster_name}@1")'
    )

    processing.run(
        "qgis:rastercalculator",
        {
            'EXPRESSION': inundation_expression,
            'LAYERS':     [base_water_layer, hand_layer],
            'OUTPUT':     inundation_raster_path
        }
    )

    if not os.path.exists(inundation_raster_path):
        print(f"  [WARNING] Raster not created for stage {stage_label}. Skipping.")
        continue

    # ------------------------------------------------------------------
    # STEP 2: Convert inundation raster pixels to points (with depth values)
    # ------------------------------------------------------------------
    inundation_points_layer = processing.run(
        "native:pixelstopoints",
        {
            'INPUT_RASTER': inundation_raster_path,
            'FIELD_NAME':   'Depth',
            'OUTPUT':       'TEMPORARY_OUTPUT'
        }
    )['OUTPUT']

    # ------------------------------------------------------------------
    # STEP 3: Load slope raster and land cover vector for pixel sampling
    # ------------------------------------------------------------------
    slope_data_provider = slope_layer.dataProvider()

    lc_layer         = QgsVectorLayer(LC_SHP_PATH, "LandCover", "ogr")
    lc_spatial_index = QgsSpatialIndex(lc_layer.getFeatures())
    lc_feature_dict  = {f.id(): f for f in lc_layer.getFeatures()}

    # ------------------------------------------------------------------
    # STEP 4: Accumulate hydraulic properties per pixel
    # ------------------------------------------------------------------
    # Initialize accumulators
    channel_area        = 0.0
    channel_bed_area    = 0.0
    floodplain_area     = 0.0
    floodplain_bed_area = 0.0

    channel_pixel_count    = 0
    floodplain_pixel_count = 0

    channel_lc_counts    = {lc_id: 0 for lc_id in lc_ids}
    floodplain_lc_counts = {lc_id: 0 for lc_id in lc_ids}

    for pixel_feature in inundation_points_layer.getFeatures():

        depth    = pixel_feature['Depth']
        geometry = pixel_feature.geometry()
        point    = geometry.asPoint()

        # Sample slope value at this pixel location
        sample_ok, sample_result = slope_data_provider.sample(point, 1)

        if sample_ok:
            if isinstance(sample_result, dict):
                slope_degrees = sample_result[1]
            elif isinstance(sample_result, list):
                slope_degrees = sample_result[0]
            else:
                slope_degrees = 0.0
        else:
            slope_degrees = 0.0

        # Convert slope from degrees to radians, handle NaN
        if math.isnan(slope_degrees):
            slope_radians = 0.0
        else:
            slope_radians = slope_degrees * (math.pi / 180.0)

        # Check if this pixel falls inside the channel mask
        is_channel_pixel = len(channel_spatial_index.intersects(geometry.boundingBox())) > 0

        # Identify the land cover class of this pixel
        overlapping_lc_ids = lc_spatial_index.intersects(geometry.boundingBox())
        lc_class_id = lc_feature_dict[overlapping_lc_ids[0]]['DN'] if overlapping_lc_ids else None

        # Compute per-pixel hydraulic area and wetted bed area
        pixel_hydraulic_area = (pixel_res ** 2 * depth) / reach_length
        pixel_bed_area       = (pixel_res ** 2) * math.sqrt(1 + (slope_radians ** 2))

        # Assign pixel to channel or floodplain
        if is_channel_pixel:
            channel_area     += pixel_hydraulic_area
            channel_bed_area += pixel_bed_area
            channel_pixel_count += 1
            if lc_class_id in channel_lc_counts:
                channel_lc_counts[lc_class_id] += 1

        elif depth >= STABILITY_DEPTH_LIMIT:
            floodplain_area     += pixel_hydraulic_area
            floodplain_bed_area += pixel_bed_area
            floodplain_pixel_count += 1
            if lc_class_id in floodplain_lc_counts:
                floodplain_lc_counts[lc_class_id] += 1

    # ------------------------------------------------------------------
    # STEP 5: Compute hydraulic radius for channel and floodplain
    # ------------------------------------------------------------------
    R_channel    = channel_area    / (channel_bed_area    / reach_length) if channel_bed_area    > 0 else 0
    R_floodplain = floodplain_area / (floodplain_bed_area / reach_length) if floodplain_bed_area > 0 else 0

    total_flow_area = channel_area + floodplain_area

    total_channel_pixels    = sum(channel_lc_counts.values())
    total_floodplain_pixels = sum(floodplain_lc_counts.values())

    # ------------------------------------------------------------------
    # STEP 6: Loop over all Manning's n combinations and compute discharge
    # ------------------------------------------------------------------
    sensitivity_results = []
    min_discharge = 999999.0

    if (total_channel_pixels + total_floodplain_pixels) > 0:

        for n_combo in all_n_combinations:

            # --- Channel discharge using area-weighted Manning's n ---
            weighted_n_channel = sum([
                channel_lc_counts[lc_id] * n_combo[lc_ids.index(lc_id)][0]
                for lc_id in lc_ids
            ])

            if total_channel_pixels > 0:
                # The weighted average n is multiplied by 1.15 to account for channel
                # sinuosity — sinuous channels have higher effective roughness due to
                # secondary flows and energy losses at bends (SI = 1.45).
                n_channel = (weighted_n_channel / total_channel_pixels) * 1.15
            else:
                n_channel = 0.040  # Default fallback roughness

            Q_channel = (1 / n_channel) * channel_area * (R_channel ** (2/3)) * (Se_adj ** 0.5)

            # --- Floodplain discharge (only active above critical depth H_crit) ---
            if stage_factor <= H_crit or total_floodplain_pixels == 0:
                Q_floodplain = 0.0
            else:
                weighted_n_fp = sum([
                    floodplain_lc_counts[lc_id] * n_combo[lc_ids.index(lc_id)][1]
                    for lc_id in lc_ids
                ])
                n_floodplain = weighted_n_fp / total_floodplain_pixels

                # Dead storage / volume reduction factor (currently set to 1.0 = no reduction)
                VRF = np.interp(stage_factor, [0.1, 3.0], [1.0, 1.0])

                Q_floodplain = (1 / n_floodplain) * floodplain_area * (R_floodplain ** (2/3)) * (Se_adj ** 0.5) * VRF

            # --- Total discharge ---
            Q_total = Q_channel + Q_floodplain

            # Store channel n values (one per land cover class) + area + Q
            channel_n_values_flat = [n_combo[i][0] for i in range(len(lc_ids))]
            sensitivity_results.append(channel_n_values_flat + [total_flow_area, Q_total])

            if Q_total < min_discharge:
                min_discharge = Q_total

        # Export sensitivity results for this stage to Excel
        column_headers    = [f"n_ID_{lc_id}" for lc_id in lc_ids] + ["Area", "Discharge_Q"]
        output_excel_path = os.path.join(output_paths["excel_files"], f"Sensitivity_{stage_label}.xlsx")

        pd.DataFrame(sensitivity_results, columns=column_headers).to_excel(
            output_excel_path,
            index=False
        )

    # Log this stage's results
    geometry_log.append({
        'Stage': stage_label,
        'Area':  round(total_flow_area, 3),
        'Min_Q': round(min_discharge, 3)
    })

    print(f"  >>> Stage {stage_label}m: Min Q = {round(min_discharge, 3)} m3/s")


# ────────────────────────
# FINAL STEP: WRITE RESULTS INTO EXISTING EXCEL SUMMARY WORKBOOK
# ────────────────────────

if os.path.exists(EXISTING_EXCEL_PATH):

    print(f"\nUpdating summary workbook: {EXISTING_EXCEL_PATH}")
    workbook  = openpyxl.load_workbook(EXISTING_EXCEL_PATH)
    worksheet = workbook[SUMMARY_SHEET_NAME]

    for entry in geometry_log:

        stage_value = float(entry['Stage'])
        target_col  = int(round(stage_value * 10)) + 2  # Column index in the workbook

        sensitivity_file = os.path.join(
            output_paths["excel_files"],
            f"Sensitivity_{entry['Stage']}.xlsx"
        )

        if not os.path.exists(sensitivity_file):
            continue

        discharge_values = pd.read_excel(sensitivity_file)["Discharge_Q"].tolist()

        # Write stage header into row 3 (handles merged cells)
        header_cell = worksheet.cell(row=3, column=target_col)

        if isinstance(header_cell, openpyxl.cell.cell.MergedCell):
            for merged_range in worksheet.merged_cells.ranges:
                if header_cell.coordinate in merged_range:
                    worksheet.cell(
                        row=merged_range.min_row,
                        column=merged_range.min_col
                    ).value = f"{entry['Stage']}m"
                    break
        else:
            header_cell.value = f"{entry['Stage']}m"

        # Write each discharge value starting from row 4
        for row_offset, q_value in enumerate(discharge_values):
            worksheet.cell(row=row_offset + 4, column=target_col).value = q_value

    workbook.save(EXISTING_EXCEL_PATH)
    print("Summary workbook updated successfully.")

else:
    print(f"\n[WARNING] Excel file not found: {EXISTING_EXCEL_PATH}")
    print("Sensitivity results were saved to individual Excel files only.")

print("\nScript completed.")
