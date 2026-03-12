"""
──────────────────────
This script performs SRC Generation with varying Manning's n roughness coefficients
for a dual-zone (channel + floodplain) hydraulic model using QGIS and HAND-based
inundation mapping.

It loops through a range of water level multipliers (stage factors), computes
cross-sectional hydraulic properties per pixel, then calculates discharge (Q)
for all possible combinations of Manning's n values per land cover class.

Results are exported to individual Excel files per stage and optionally written
into an existing summary workbook.

Requirements:
    - QGIS Python environment (qgis.core, processing)
    - NumPy, Pandas, OpenPyXL, itertools, math

Author: [Your Name]
Date: [Date]
"""

# ────────────────────────
# IMPORTS
# ────────────────────────
import os
import math
import itertools
import numpy as np
import pandas as pd
import openpyxl
from qgis.core import (QgsProject, QgsVectorLayer, QgsSpatialIndex)
import processing

# ────────────────────────────────────────────────────────────────
#  CONFIGURATION & INPUTS
# ────────────────────────────────────────────────────────────────

# Base directory
ROOT_DIR = r"C:\Users\acer\Downloads\aQGIS Content\Official Sample\CDO_NewMethodology\Test Web Length"

# Input Files
LC_SHP_PATH       = os.path.join(ROOT_DIR, "Input 2", "LandCover_Vector_3984,6m.shp")
CHANNEL_MASK_PATH = os.path.join(ROOT_DIR, "FINAL CODE", "Smaller Channel", "SmallerChannel_Vector.shp")
EXISTING_EXCEL    = os.path.join(ROOT_DIR, "QGIS-Calculations-San-Isidro.xlsx")

# Output Folders
OUT_RASTER = os.path.join(ROOT_DIR, "Output", "Flood Rasters")
OUT_EXCEL  = os.path.join(ROOT_DIR, "Output", "Excel Files")

# Ensure folders exist
for p in [OUT_RASTER, OUT_EXCEL]:
    os.makedirs(p, exist_ok=True)

# Hydraulic Parameters
REACH_LENGTH    = 3984.6
PIXEL_RES       = 5.0
VALLEY_SLOPE    = (42.80544281 - 29.117033) / 2749.334
H_CRIT          = 1.8
STABILITY_LIMIT = 0.3

# Layer Names (Must match QGIS Table of Contents)
LAYER_WATER = "PA_WL_Cabula_10_FINAL"
LAYER_HAND  = "HAND_EleveAbvStream_BestSubcatchmentClipped"
LAYER_SLOPE = "Slope_Degrees_NewDEM"

# ────────────────────────────────────────────────────────────────
#  ROUGHNESS CONFIGURATION (MANNING'S N)
#  Format: {ID: {'name': Type, 'chan': [start, end, step], ...}}
# ────────────────────────────────────────────────────────────────
n_configs = {
    1:  {'name': 'Water',      'chan': [0.025, 0.080, 0.005], 'fp': [0.025, 0.080, 0.005]},
    11: {'name': 'Rangelands', 'chan': [0.030, 0.070, 0.001], 'fp': [0.020, 0.040, 0.001]},
    7:  {'name': 'Built Area', 'chan': [0.065, 0.120, 0.050], 'fp': [0.065, 0.120, 0.050]},
    5:  {'name': 'Crops',      'chan': [0.025, 0.055, 0.010], 'fp': [0.025, 0.055, 0.010]},
    2:  {'name': 'Trees',      'chan': [0.120, 0.300, 0.005], 'fp': [0.120, 0.300, 0.005]}
}

# ────────────────────────────────────────────────────────────────
#  INITIALIZATION
# ────────────────────────────────────────────────────────────────

def get_layer(name):
    layers = QgsProject.instance().mapLayersByName(name)
    return layers[0] if layers else None

base_layer  = get_layer(LAYER_WATER)
hand_layer  = get_layer(LAYER_HAND)
slope_layer = get_layer(LAYER_SLOPE)

# Load Vectors and Spatial Indices
mask_layer    = QgsVectorLayer(CHANNEL_MASK_PATH, "Channel_Mask", "ogr")
channel_index = QgsSpatialIndex(mask_layer.getFeatures())

lc_layer    = QgsVectorLayer(LC_SHP_PATH, "LC", "ogr")
lc_index    = QgsSpatialIndex(lc_layer.getFeatures())
lc_features = {f.id(): f for f in lc_layer.getFeatures()}

# Generate Parameter Combinations
keys        = list(n_configs.keys())
param_lists = []

for k in keys:
    cfg    = n_configs[k]
    c_steps = np.linspace(*cfg['chan'], int(round((cfg['chan'][1] - cfg['chan'][0]) / cfg['chan'][2])) + 1)
    f_steps = np.linspace(*cfg['fp'],  int(round((cfg['fp'][1]   - cfg['fp'][0])   / cfg['fp'][2]))  + 1)
    pairs   = list(zip(np.round(c_steps, 3), np.round(f_steps, 3)))
    param_lists.append(pairs)

combinations = list(itertools.product(*param_lists))
geometry_log = []

# ────────────────────────────────────────────────────────────────
#  MAIN PROCESSING LOOP
# ────────────────────────────────────────────────────────────────
# Loop through water levels from 0.1m to 3.0m
for i in range(1, 31):
    f_val = i / 10.0
    f_str = f"{f_val:.1f}"

    # 1. Raster Calculation (Water Depth)
    out_tif = os.path.join(OUT_RASTER, f"Inun_{f_str}.tif")
    expr = (
        f'(("{LAYER_WATER}@1" * {f_val}) - "{LAYER_HAND}@1" > 0) * '
        f'(("{LAYER_WATER}@1" * {f_val}) - "{LAYER_HAND}@1")'
    )

    processing.run("qgis:rastercalculator", {
        'EXPRESSION': expr,
        'LAYERS':     [base_layer, hand_layer],
        'OUTPUT':     out_tif
    })

    if not os.path.exists(out_tif):
        continue

    # 2. Convert Pixels to Points
    pts_layer = processing.run("native:pixelstopoints", {
        'INPUT_RASTER': out_tif,
        'FIELD_NAME':   'Depth',
        'OUTPUT':       'TEMPORARY_OUTPUT'
    })['OUTPUT']

    slope_prov = slope_layer.dataProvider()

    # Reset accumulators
    chan_A = 0.0
    fp_A   = 0.0
    chan_BedArea = {k: 0.0 for k in keys}  # Equivalent to P_i (Wetted Perimeter)
    fp_BedArea   = {k: 0.0 for k in keys}

    # 3. Geometric Analysis per Pixel
    for feat in pts_layer.getFeatures():
        depth = feat['Depth']
        geom  = feat.geometry()

        # Sample slope
        res, sample = slope_prov.sample(geom.asPoint(), 1)
        slope_deg   = sample[1] if res else 0
        slope_rad   = math.radians(slope_deg if not math.isnan(slope_deg) else 0)

        # Check location (Channel vs Floodplain)
        is_chan = len(channel_index.intersects(geom.boundingBox())) > 0

        # Identify Land Cover ID
        intersect_ids = lc_index.intersects(geom.boundingBox())
        lc_id = lc_features[intersect_ids[0]]['DN'] if intersect_ids else None

        # Calculate Geometry
        a_p = (PIXEL_RES ** 2 * depth) / REACH_LENGTH

        # Wetted Perimeter (Corrected for slope to get true bed area)
        cos_val = math.cos(slope_rad)
        if abs(cos_val) < 1e-6:
            cos_val = 1e-6

        bed_p = (PIXEL_RES ** 2) / cos_val

        # Aggregate Data
        if is_chan:
            chan_A += a_p
            if lc_id in chan_BedArea:
                chan_BedArea[lc_id] += bed_p
        elif depth >= STABILITY_LIMIT:
            fp_A += a_p
            if lc_id in fp_BedArea:
                fp_BedArea[lc_id] += bed_p

    # 4. Hydraulic Calculations
    total_chan_P = sum(chan_BedArea.values())
    total_fp_P   = sum(fp_BedArea.values())

    # Hydraulic Radius (R = A / P_total)
    R_chan = chan_A / (total_chan_P / REACH_LENGTH) if total_chan_P > 0 else 0
    R_fp   = fp_A  / (total_fp_P  / REACH_LENGTH) if total_fp_P  > 0 else 0

    current_A = chan_A + fp_A
    min_Q     = 999999.0
    results   = []

    SI     = 1.45
    Se_adj = VALLEY_SLOPE / SI

    # 5. Sensitivity Analysis (Iterate n-value combinations)
    if (total_chan_P + total_fp_P) > 0:
        for combo in combinations:

            # ──────────────────────────────────────────────────────────
            # METHODOLOGY: Einstein-Horton Equation
            # Used to calculate composite roughness (n_c) for
            # heterogeneous beds.
            # Formula: n_c = [ Sum(P_i * n_i^1.5) / P_total ]^(2/3)
            # ──────────────────────────────────────────────────────────

            # --- Channel Composite Roughness ---
            if total_chan_P > 0:
                # Weighting n-values by their Wetted Perimeter (P_i)
                w_n_chan = sum([
                    (chan_BedArea[k] / total_chan_P) * (combo[keys.index(k)][0] ** 1.5)
                    for k in keys
                ])
                n_chan = w_n_chan ** (2 / 3)
            else:
                n_chan = 0.040

            # Calculate Channel Discharge (Manning's)
            Q_chan = (1 / n_chan) * chan_A * (R_chan ** (2 / 3)) * (Se_adj ** 0.5)

            # --- Floodplain Composite Roughness ---
            if f_val <= H_CRIT or total_fp_P == 0:
                Q_fp = 0
            else:
                # Apply Einstein-Horton weighting for Floodplain Land Cover
                w_n_fp = sum([
                    (fp_BedArea[k] / total_fp_P) * (combo[keys.index(k)][1] ** 1.5)
                    for k in keys
                ])
                n_fp = w_n_fp ** (2 / 3)

                # Volume Reduction Factor (VRF)
                VRF  = np.interp(f_val, [0.1, 3.0], [1.0, 1.0])
                Q_fp = (1 / n_fp) * fp_A * (R_fp ** (2 / 3)) * (Se_adj ** 0.5) * VRF

            Q_total = Q_chan + Q_fp

            flat_n = [c[0] for c in combo]
            results.append(flat_n + [current_A, Q_total])
            if Q_total < min_Q:
                min_Q = Q_total

        # Export Stage Sensitivity
        df = pd.DataFrame(results, columns=[f"n_ID_{k}" for k in keys] + ["Area", "Discharge"])
        df.to_excel(os.path.join(OUT_EXCEL, f"Sensitivity_{f_str}.xlsx"), index=False)

    geometry_log.append({'Stage': f_str, 'Area': current_A, 'Min_Q': min_Q})
    print(f"   > Stage {f_str}m Processed. Min Q: {round(min_Q, 3)}")

# ────────────────────────────────────────────────────────────────
#  SUMMARY EXPORT
# ────────────────────────────────────────────────────────────────
if os.path.exists(EXISTING_EXCEL):
    wb = openpyxl.load_workbook(EXISTING_EXCEL)
    ws = wb["Cabula Curve List"]

    for entry in geometry_log:
        stage   = entry['Stage']
        col_idx = int(round(float(stage) * 10)) + 2

        sens_path = os.path.join(OUT_EXCEL, f"Sensitivity_{stage}.xlsx")
        if os.path.exists(sens_path):
            qs = pd.read_excel(sens_path)["Discharge"].tolist()

            cell       = ws.cell(row=3, column=col_idx)
            cell.value = f"{stage}m"

            for idx, q in enumerate(qs):
                ws.cell(row=idx + 4, column=col_idx).value = q

    wb.save(EXISTING_EXCEL)
    print("Optimization Complete.")
